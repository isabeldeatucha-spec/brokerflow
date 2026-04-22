"""Admin & Ops — LangGraph graph definition.

Flow:
  load_brand_context
       ↓ [miss/stale] ──→ END
       ↓ [ok]
  rule_based_autofill
       ↓
  llm_inference_pass
       ↓
  flag_gaps
       ↓
  generate_filled_xlsx
       ↓
      END
"""
from __future__ import annotations

import json
import os
import re
import uuid
from datetime import datetime, timezone
from typing import Any

import anthropic
from langgraph.graph import StateGraph, END

from memory import memory, _get_client, get_config, store_new_item_form
from state import AdminOpsFormFillState
from agents.admin_ops.skills.wfm_form_schema import WFM_FORM_FIELDS, FIELD_SECTIONS


HANDOFF_FRESHNESS_SECONDS = 30 * 24 * 60 * 60  # 30 days
MODEL = "claude-haiku-4-5-20251001"
TEMPLATE_PATH = os.path.join(
    os.path.dirname(__file__), "skills", "2018_WFM_Global_New_Item_Setup_Form.xlsx"
)

# Stable section ordering for gap sorting
SECTION_ORDER = list(FIELD_SECTIONS.keys())

_SUGGESTED_ACTIONS: dict[str, str] = {
    "text":       "Enter the text value for this field.",
    "number":     "Enter a numeric value.",
    "currency":   "Enter the dollar amount (e.g. 12.99).",
    "yes_no":     "Enter Y or N.",
    "date":       "Enter the date (MM/YYYY or MM/DD/YYYY).",
    "dimensions": "Enter the measurement in inches.",
}

_LLM_INFERENCE_PROMPT = """
You are filling out a Whole Foods Market new item setup form for a food brand.
Use the brand context below to infer values for these four fields only.

Brand context:
{context_json}

Fields to infer:
- family: The product family or line name (e.g. "Original", "Classic Collection", "Everyday"). \
If there is only one product line, use the brand name.
- description: A single polished product description line (max 100 chars) suitable for a buyer form. \
Rewrite the broker_brief into one tight sentence if it is too long.
- gm_pct: Estimated gross margin percentage (number, e.g. 42.5). \
Natural/specialty food typical GM% by category: snack bars 40-45%, beverages 35-42%, \
condiments 42-50%, dairy 38-44%, supplements 50-60%. Use SRP and category to estimate if known.
- is_line_extension: "Y" if signals show multiple SKUs / flavors / variants, otherwise "N".

Return ONLY valid JSON — no prose, no markdown fences:
{{
  "family":           {{"value": "...", "confidence": "medium", "reasoning": "..."}},
  "description":      {{"value": "...", "confidence": "high",   "reasoning": "..."}},
  "gm_pct":           {{"value": 42.5,  "confidence": "low",    "reasoning": "..."}},
  "is_line_extension":{{"value": "N",   "confidence": "medium", "reasoning": "..."}}
}}
"""


# ── Helpers ───────────────────────────────────────────────────────────────────

def _strip_fences(text: str) -> str:
    text = text.strip()
    if text.startswith("```"):
        parts = text.split("```")
        text = parts[1] if len(parts) > 1 else text
        if text.startswith("json"):
            text = text[4:]
    return text.strip()


def _slugify(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")


def _cert_flag(certs: list[str] | None, *keywords: str) -> str | None:
    """Return 'Y' if any cert string contains a keyword, 'N' if certs exist but no match."""
    if not certs:
        return None
    lower = " ".join(c.lower() for c in certs)
    return "Y" if any(kw in lower for kw in keywords) else "N"


def _get_flagship_sku(brand_record: dict) -> dict:
    """Return the flagship SKU dict from the products array, or empty dict if none."""
    products = brand_record.get("products") or []
    for p in products:
        if isinstance(p, dict) and p.get("is_flagship"):
            return p
    return products[0] if products else {}


# ── Nodes ─────────────────────────────────────────────────────────────────────

def load_brand_context(state: AdminOpsFormFillState) -> dict:
    """Load brand context from Supabase. Tries brands table first, falls back to brand_evaluations."""
    brand_name = state["brand_name"]
    new_errors: list[str] = []

    try:
        client = _get_client()

        # Try canonical brands table first (populated by onboarding agent)
        brands_res = (
            client.table("brands")
            .select("*")
            .ilike("brand_name", brand_name)
            .limit(1)
            .execute()
        )
        if brands_res.data:
            canonical = brands_res.data[0]
            merged_row = {
                "brand_name": canonical.get("brand_name", brand_name),
                "category": canonical.get("category", ""),
                "extracted_fields": canonical,
                "score": canonical.get("completeness_pct", 0),
                "verdict": "broker_ready",
                "broker_brief": canonical.get("brand_story", ""),
                "founder_name": canonical.get("founder_name", ""),
                "founder_email": canonical.get("founder_email", ""),
                "evaluated_at": canonical.get("last_verified_at", ""),
            }
            res_data = [merged_row]
        else:
            # -- Brand Scout evaluation (required) ---------------------------------
            res = (
                client.table("brand_evaluations")
                .select("*")
                .ilike("brand_name", brand_name)
                .order("evaluated_at", desc=True)
                .limit(1)
                .execute()
            )
            if not res.data:
                return {
                    "handoff_status": "miss",
                    "handoff_error": f"No Brand Scout evaluation found for {brand_name!r}.",
                    "scout_context": {},
                    "pitcher_context": {},
                    "form_schema": WFM_FORM_FIELDS,
                    "artifact_errors": [f"load: no Brand Scout row for {brand_name!r}"],
                }
            res_data = res.data

        row = res_data[0]
        evaluated_at = row.get("evaluated_at", "")
        if evaluated_at:
            try:
                ts = datetime.fromisoformat(evaluated_at.replace("Z", "+00:00"))
                age = (datetime.now(timezone.utc) - ts).total_seconds()
                if age > HANDOFF_FRESHNESS_SECONDS:
                    return {
                        "handoff_status": "stale",
                        "handoff_error": f"Scout evaluation is {int(age / 3600)}h old (>{HANDOFF_FRESHNESS_SECONDS // 3600}h limit).",
                        "scout_context": row,
                        "pitcher_context": {},
                        "form_schema": WFM_FORM_FIELDS,
                        "artifact_errors": [f"load: stale Scout row ({int(age / 3600)}h old)"],
                    }
            except ValueError:
                pass

        # -- Retailer Pitcher context (optional) -------------------------------
        pitcher_ctx: dict = {}
        try:
            pitch_res = (
                client.table("retailer_pitches")
                .select("*")
                .ilike("brand_name", brand_name)
                .order("created_at", desc=True)
                .limit(1)
                .execute()
            )
            if pitch_res.data:
                pitcher_ctx = pitch_res.data[0]
        except Exception as exc:  # noqa: BLE001
            new_errors.append(f"load_pitcher: {type(exc).__name__}: {exc}")

        return {
            "handoff_status": "ok",
            "handoff_error": None,
            "scout_context": row,
            "pitcher_context": pitcher_ctx,
            "form_schema": WFM_FORM_FIELDS,
            "artifact_errors": new_errors,
        }

    except Exception as exc:  # noqa: BLE001
        return {
            "handoff_status": "miss",
            "handoff_error": f"{type(exc).__name__}: {exc}",
            "scout_context": {},
            "pitcher_context": {},
            "form_schema": WFM_FORM_FIELDS,
            "artifact_errors": [f"load: {type(exc).__name__}: {exc}"],
        }


def rule_based_autofill(state: AdminOpsFormFillState) -> dict:
    """Populate form fields deterministically from scout_context. No LLM call."""
    ctx: dict = state.get("scout_context", {})
    key_signals: dict = ctx.get("key_signals", {}) or {}
    extracted: dict = key_signals.get("extracted_fields", {}) or {}
    certs: list[str] | None = extracted.get("certifications")

    filled: dict[str, Any] = dict(state.get("filled_fields", {}))
    confidence: dict[str, str] = dict(state.get("field_confidence", {}))
    sources: dict[str, str] = dict(state.get("field_sources", {}))

    def _set(field_id: str, value: Any, conf: str, source: str) -> None:
        if value is None or value == "":
            return
        filled[field_id] = value
        confidence[field_id] = conf
        sources[field_id] = source

    # -- Flagship SKU (prefer brands table products array) --------------------
    # When ctx comes from the brands table, extracted_fields IS the canonical record
    brands_record: dict = ctx.get("extracted_fields", {}) or {}
    flagship = _get_flagship_sku(brands_record) or _get_flagship_sku(ctx)
    if not certs:
        certs = brands_record.get("certifications")

    # -- Direct copies --------------------------------------------------------
    _set("brand",               ctx.get("brand_name"),   "high", "Brand Scout · brand_name")
    _set("category",            ctx.get("category"),     "high", "Brand Scout · category")
    _set("manufacturer_company", ctx.get("brand_name"), "high", "Brand Scout · brand_name")
    _set("vendor_contact_name", ctx.get("founder_name"), "high", "Brand Scout · founder_name")
    _set("vendor_contact_email", ctx.get("founder_email"), "high", "Brand Scout · founder_email")

    # -- Flagship SKU fields: real unit-level data, not brand-level ranges ----
    if flagship:
        _set("upc",         flagship.get("upc"),             "high", "brands · flagship_sku.upc")
        _set("case_pack",   flagship.get("case_pack"),        "high", "brands · flagship_sku.case_pack")
        _set("shelf_life",  flagship.get("shelf_life_days"),  "high", "brands · flagship_sku.shelf_life_days")
        _set("net_weight",  flagship.get("net_weight"),       "high", "brands · flagship_sku.net_weight")
        _set("storage_temp", flagship.get("storage_temp"),   "high", "brands · flagship_sku.storage_temp")
        wc = flagship.get("wholesale_cost")
        if wc is not None:
            _set("wholesale_cost", wc, "high", "brands · flagship_sku.wholesale_cost")
        msrp = flagship.get("msrp")
        if msrp is not None:
            _set("srp", msrp, "high", "brands · flagship_sku.msrp")

    # -- Description: first 200 chars of broker_brief -------------------------
    broker_brief: str = ctx.get("broker_brief", "") or ""
    if broker_brief:
        _set("description", broker_brief[:200].strip(), "high", "Brand Scout · broker_brief")

    # -- SRP: flagship msrp already set above; fall back to extracted signals --
    if "srp" not in filled:
        srp = extracted.get("srp_hero") or extracted.get("srp_min")
        if srp is None:
            srp = key_signals.get("srp_hero") or key_signals.get("srp_min")
        if srp is not None:
            _set("srp", srp, "high", "Brand Scout · extracted_fields.srp_hero")

    # -- Certification flags (yes_no) — medium confidence because inferred ----
    usda = _cert_flag(certs, "usda organic", "organic")
    _set("usda_organic", usda, "medium", "Brand Scout · extracted_fields.certifications")

    ngmo = _cert_flag(certs, "non-gmo", "non gmo", "non gmo project")
    _set("non_gmo_verified", ngmo, "medium", "Brand Scout · extracted_fields.certifications")

    ft = _cert_flag(certs, "fair trade")
    _set("fair_trade", ft, "medium", "Brand Scout · extracted_fields.certifications")

    gf = _cert_flag(certs, "gluten free", "gluten-free", "certified gluten")
    _set("third_party_gluten_free", gf, "medium", "Brand Scout · extracted_fields.certifications")

    biodynamic = _cert_flag(certs, "biodynamic", "demeter")
    _set("demeter_biodynamic", biodynamic, "medium", "Brand Scout · extracted_fields.certifications")

    return {
        "filled_fields": filled,
        "field_confidence": confidence,
        "field_sources": sources,
    }


def llm_inference_pass(state: AdminOpsFormFillState) -> dict:
    """Ask Haiku to infer family, description (rewrite), gm_pct, and is_line_extension."""
    ctx: dict = state.get("scout_context", {})
    new_errors: list[str] = []

    context_payload = {
        "brand_name":    ctx.get("brand_name", state["brand_name"]),
        "category":      ctx.get("category", ""),
        "broker_brief":  ctx.get("broker_brief", ""),
        "score":         ctx.get("score", None),
        "key_signals":   ctx.get("key_signals", {}),
        "score_breakdown": ctx.get("score_breakdown", {}),
        "srp_already_filled": state.get("filled_fields", {}).get("srp"),
    }

    prompt = _LLM_INFERENCE_PROMPT.format(
        context_json=json.dumps(context_payload, indent=2)[:4000]
    )

    try:
        client = anthropic.Anthropic(api_key=os.environ["ANTHROPIC_API_KEY"])
        resp = client.messages.create(
            model=MODEL,
            max_tokens=1000,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = resp.content[0].text if resp.content else ""
        data: dict[str, dict] = json.loads(_strip_fences(raw))
    except json.JSONDecodeError as exc:
        new_errors.append(f"llm_inference: JSON parse error: {exc}")
        return {"artifact_errors": new_errors}
    except Exception as exc:  # noqa: BLE001
        new_errors.append(f"llm_inference: {type(exc).__name__}: {exc}")
        return {"artifact_errors": new_errors}

    filled: dict[str, Any] = dict(state.get("filled_fields", {}))
    confidence: dict[str, str] = dict(state.get("field_confidence", {}))
    sources: dict[str, str] = dict(state.get("field_sources", {}))

    for field_id, payload in data.items():
        if not isinstance(payload, dict):
            continue
        value = payload.get("value")
        if value is None or value == "":
            continue
        # Don't overwrite a high-confidence rule-based value
        if confidence.get(field_id) == "high":
            continue
        filled[field_id] = value
        confidence[field_id] = payload.get("confidence", "medium")
        reasoning = payload.get("reasoning", "")
        sources[field_id] = f"LLM inference · {reasoning[:80]}" if reasoning else "LLM inference"

    return {
        "filled_fields": filled,
        "field_confidence": confidence,
        "field_sources": sources,
        "artifact_errors": new_errors,
    }


def flag_gaps(state: AdminOpsFormFillState) -> dict:
    """Collect fields the agent couldn't fill. Sort required gaps first."""
    filled: dict[str, Any] = state.get("filled_fields", {})
    confidence: dict[str, str] = state.get("field_confidence", {})

    gaps: list[dict] = []
    for field in WFM_FORM_FIELDS:
        fid = field["id"]
        missing = fid not in filled or confidence.get(fid) == "missing"
        if not missing:
            continue

        if field["required"]:
            reason = "required — broker must provide"
            action = _SUGGESTED_ACTIONS.get(field["type"], "Enter the value for this field.")
        else:
            reason = "optional — fill if available"
            action = _SUGGESTED_ACTIONS.get(field["type"], "Enter the value if known.")

        gaps.append({
            "field_id":        fid,
            "label":           field["label"],
            "section":         field["section"],
            "required":        field["required"],
            "reason":          reason,
            "suggested_action": action,
        })

    # Sort: required first, then by section order, then by field list position
    field_positions = {f["id"]: i for i, f in enumerate(WFM_FORM_FIELDS)}
    gaps.sort(key=lambda g: (
        0 if g["required"] else 1,
        SECTION_ORDER.index(g["section"]) if g["section"] in SECTION_ORDER else 99,
        field_positions.get(g["field_id"], 999),
    ))

    return {"gaps": gaps}


def generate_filled_xlsx(state: AdminOpsFormFillState) -> dict:
    """Write filled values into the WFM Excel template and save to /tmp."""
    new_errors: list[str] = []
    brand_slug = _slugify(state["brand_name"])
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    out_path = f"/tmp/sedge_wfm_{brand_slug}_{timestamp}.xlsx"

    try:
        import openpyxl  # imported here to avoid hard dep at module load

        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        ws = wb["New Item Setup Form"]

        # Clear every schema column in row 19 so template sample data doesn't
        # bleed through into un-filled cells (includes hyperlinks on email cells).
        for field in WFM_FORM_FIELDS:
            cell = ws[f"{field['excel_column']}19"]
            cell.value = None
            cell.hyperlink = None

        filled: dict[str, Any] = state.get("filled_fields", {})
        schema_by_id = {f["id"]: f for f in WFM_FORM_FIELDS}

        for field_id, value in filled.items():
            field = schema_by_id.get(field_id)
            if not field:
                continue
            cell_addr = f"{field['excel_column']}19"
            ws[cell_addr] = value

        wb.save(out_path)
        status = "partial" if state.get("gaps") else "ok"

        store_new_item_form(
            brand_name=       state["brand_name"],
            retailer=         state.get("retailer", "whole_foods"),
            filled_fields=    state.get("filled_fields", {}),
            field_confidence= state.get("field_confidence", {}),
            field_sources=    state.get("field_sources", {}),
            gaps=             state.get("gaps", []),
            output_xlsx_path= out_path,
            output_status=    status,
        )

    except FileNotFoundError:
        msg = f"WFM template not found at {TEMPLATE_PATH}"
        new_errors.append(f"xlsx: {msg}")
        return {
            "output_xlsx_path": "",
            "output_status": "failed",
            "artifact_errors": new_errors,
        }
    except Exception as exc:  # noqa: BLE001
        new_errors.append(f"xlsx: {type(exc).__name__}: {exc}")
        return {
            "output_xlsx_path": "",
            "output_status": "failed",
            "artifact_errors": new_errors,
        }

    return {
        "output_xlsx_path": out_path,
        "output_status": status,
        "artifact_errors": new_errors,
    }


# ── Routing ───────────────────────────────────────────────────────────────────

def _route_after_load(state: AdminOpsFormFillState) -> str:
    return "ok" if state["handoff_status"] == "ok" else "stop"


# ── Graph ─────────────────────────────────────────────────────────────────────

def build_graph():
    g = StateGraph(AdminOpsFormFillState)

    g.add_node("load_brand_context",   load_brand_context)
    g.add_node("rule_based_autofill",  rule_based_autofill)
    g.add_node("llm_inference_pass",   llm_inference_pass)
    g.add_node("flag_gaps",            flag_gaps)
    g.add_node("generate_filled_xlsx", generate_filled_xlsx)

    g.set_entry_point("load_brand_context")

    g.add_conditional_edges(
        "load_brand_context",
        _route_after_load,
        {"ok": "rule_based_autofill", "stop": END},
    )

    g.add_edge("rule_based_autofill",  "llm_inference_pass")
    g.add_edge("llm_inference_pass",   "flag_gaps")
    g.add_edge("flag_gaps",            "generate_filled_xlsx")
    g.add_edge("generate_filled_xlsx", END)

    return g.compile(checkpointer=memory)


graph = build_graph()


# ── Convenience runner ────────────────────────────────────────────────────────

def run_admin_ops(brand_name: str, retailer: str = "whole_foods") -> dict:
    """Run the Admin & Ops form-fill graph for one brand and return final state."""
    thread_id = str(uuid.uuid4())
    config = get_config(thread_id)

    initial: AdminOpsFormFillState = {
        "brand_name":       brand_name,
        "retailer":         retailer,
        "scout_context":    {},
        "pitcher_context":  {},
        "handoff_status":   "",
        "handoff_error":    None,
        "form_schema":      [],
        "filled_fields":    {},
        "field_confidence": {},
        "field_sources":    {},
        "gaps":             [],
        "output_xlsx_path": "",
        "output_status":    "",
        "artifact_errors":  [],
        "approved":         None,
        "rejection_reason": None,
    }

    for _ in graph.stream(initial, config=config, stream_mode="updates"):
        pass

    return graph.get_state(config).values
